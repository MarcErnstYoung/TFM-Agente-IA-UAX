#!/usr/bin/env python3
# predictor.py
# Uso: python predictor.py data_TPFP.csv
# Requisitos: preprocessor.joblib y Model_V1.h5 en la misma carpeta.

import sys
import os
import joblib
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from scipy import sparse as sp
from sklearn.metrics import (
    precision_recall_curve,
    average_precision_score,
    classification_report,
    confusion_matrix
)
import argparse
import warnings

# Para cargar el modelo Keras sin necesidad de recompilar la arquitectura (evitamos dependencias de loss/metrics)
import tensorflow as tf
from tensorflow.keras.models import load_model # type: ignore

warnings.filterwarnings("ignore")

def load_preprocessor(path="preprocessor.joblib"):
    if not os.path.exists(path):
        raise FileNotFoundError(f"No se encontró el preprocessor en: {path}")
    preproc = joblib.load(path)
    print(f"[OK] Preprocessor cargado desde: {path}")
    return preproc

def load_keras_model(path="Model_V1.h5"):
    if not os.path.exists(path):
        raise FileNotFoundError(f"No se encontró el modelo Keras en: {path}")
    # Cargamos sin compilar para evitar problemas con funciones de pérdida/ métricas personalizadas
    model = load_model(path, compile=False)
    print(f"[OK] Modelo Keras cargado desde: {path} (compile=False)")
    return model

def prepare_dataframe(csv_path):
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"No se encontró el CSV: {csv_path}")

    df = pd.read_csv(csv_path, low_memory=False)
    df.columns = df.columns.str.strip()

    # --- FIX 1: rellenar NaN globalmente ---
    df = df.replace({np.nan: ""})

    # --- FIX 2: asegurar strings en columnas tipo objeto ---
    obj_cols = df.select_dtypes(include=["object"]).columns
    df[obj_cols] = df[obj_cols].astype(str)

    # Normalización específica que ya tenías
    if "dest_port" in df.columns:
        df["dest_port"] = (
            df["dest_port"]
            .astype(str)
            .str.replace(r"[\[\],]", " ", regex=True)
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
        )

    return df

def transform_with_preprocessor(preproc, df):
    # preproc.transform espera el DataFrame con las columnas que usó en fit
    try:
        X = preproc.transform(df)
    except Exception as e:
        # Mensaje diagnóstico: a menudo es por columnas faltantes o tipos inesperados
        raise RuntimeError(f"Error al transformar con el preprocessor: {e}")
    return X

def to_model_input(X):
    # Si es sparse, convertir a dense (el modelo espera dense float32)
    if sp.issparse(X):
        X_in = X.toarray()
    else:
        X_in = np.asarray(X)
    X_in = X_in.astype("float32")
    return X_in

def compute_best_threshold(y_true, y_prob):
    prec, rec, th = precision_recall_curve(y_true, y_prob)
    # precision_recall_curve returns len(th) = len(prec)-1
    if len(th) == 0:
        return 0.5  # fallback
    f1 = 2 * prec[:-1] * rec[:-1] / (prec[:-1] + rec[:-1] + 1e-12)
    best_idx = int(np.nanargmax(f1))
    return float(th[best_idx])

def print_basic_stats(df, y_true, y_prob, y_pred):
    print("\n=== Resumen de resultados ===")
    total = len(y_prob)
    print(f"Registros procesados: {total}")
    if y_true is not None:
        positives = int((y_true == 1).sum())
        negs = int((y_true == 0).sum())
        print(f"Etiquetas reales: {positives} positivas, {negs} negativas")
    print(f"Prob. (media): {y_prob.mean():.4f}  |  Prob. (mediana): {np.median(y_prob):.4f}")
    if y_true is not None:
        pr_auc = average_precision_score(y_true, y_prob)
        print(f"PR AUC (average precision): {pr_auc:.4f}")
        print("\nClassification report (umbral usado para predicción):")
        print(classification_report(y_true, y_pred, digits=4))
        print("\nConfusion matrix (orden: [neg, pos]):")
        print(confusion_matrix(y_true, y_pred))
    else:
        print("No se proporcionó 'disposition_label' → no se calcularon métricas de referencia.")
    print("=============================\n")

def main(args):
    csv_path = args.csv
    preproc_path = args.preprocessor
    model_path = args.model
    output_csv = args.output
    default_threshold = args.default_threshold

    # 1) Cargar preprocessor y modelo
    preproc = load_preprocessor(preproc_path)
    model = load_keras_model(model_path)

    # 2) Cargar CSV
    df = prepare_dataframe(csv_path)
    print(f"[OK] CSV cargado: {csv_path} -> {df.shape[0]} filas, {df.shape[1]} columnas")

    # 3) Si hay columna disposition_label la extraemos (para evaluación)
    y = None
    if "disposition_label" in df.columns:
        # normalizamos como en entrenamiento
        y_series = (df["disposition_label"].astype(str).str.strip().str.upper())
        mask = y_series.str.startswith(("TP", "FP"))
        # Si hay filas no TP/FP las mantenemos pero convertimos a nan en y (evitamos crash)
        if not mask.all():
            print("[WARN] Algunas filas tienen disposition_label distinto a TP/FP. Se considerarán como 'no etiqueta' para evaluación.")
        # Only mark y for rows that startwith TP/FP
        y = np.where(y_series.str.startswith("TP"), 1, 0)
        # Si hay rows con otros valores, y igualará 0/1 según startswith; user should supply TP/FP.
    else:
        y = None

    # 4) Transformar
    try:
        X_trans = transform_with_preprocessor(preproc, df)
    except Exception as e:
        print(f"[ERROR] fallo en el preprocesado: {e}")
        sys.exit(1)

    # 5) Convertir a input del modelo
    X_in = to_model_input(X_trans)
    print(f"[OK] Datos transformados. Shape para el modelo: {X_in.shape}")

    # 6) Inferencia (probabilidades)
    try:
        y_prob = model.predict(X_in, batch_size=256).ravel()
    except Exception as e:
        # En caso de que el modelo devuelva shape (n,1) o similar, normalizamos
        preds = model.predict(X_in, batch_size=256)
        y_prob = np.asarray(preds).reshape(-1)
    print("[OK] Inferencia completa.")

    # 7) Determinar umbral
    used_threshold = default_threshold
    if y is not None:
        try:
            used_threshold = compute_best_threshold(y, y_prob)
            print(f"[INFO] Umbral óptimo calculado (max F1) = {used_threshold:.4f}")
        except Exception as e:
            print(f"[WARN] No se pudo calcular umbral óptimo, usando default {default_threshold}: {e}")

    else:
        print(f"[INFO] No hay labels reales. Usando umbral por defecto = {default_threshold}")

    y_pred = (y_prob >= used_threshold).astype(int)

    # 8) Impresión de resultados por terminal (resumen + top ejemplos)
    print_basic_stats(df, y, y_prob, y_pred)

    # Mostrar primeras 10 filas con probabilidades y predicción
    display_cols = []
    # intentamos mostrar columnas identificadoras si existen
    for c in ["orig_client", "signature", "orig_rule_title", "subject", "sender", "recipient", "user", "hostname", "url"]:
        if c in df.columns and len(display_cols) < 3:
            display_cols.append(c)
    head_df = df.loc[:, display_cols].copy() if display_cols else df.iloc[:, :0].copy()
    head_df = head_df.reset_index(drop=True)
    head_df["probability"] = y_prob
    head_df["predicted_label"] = y_pred
    if "disposition_label" in df.columns:
        head_df["true_label"] = y
    print("Primeras 10 filas (columnas relevantes + probabilidad + predicción):")
    print(head_df.head(10).to_string(index=False))

    # 9) Guardar archivo de salida con probabilidades y predicción
    out = df.copy().reset_index(drop=True)
    out["pred_probability"] = y_prob
    out["pred_label"] = y_pred
    if "disposition_label" in df.columns:
        out["true_label"] = y

    # Nombre del Excel
    excel_output = output_csv.replace(".csv", ".xlsx")

    # Guardamos primero con pandas para volcar datos
    with pd.ExcelWriter(excel_output, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="Predicciones", index=False)

    # Abrimos con openpyxl para crear la tabla
    wb = openpyxl.load_workbook(excel_output)
    ws = wb["Predicciones"]

    # Definir rango de tabla (desde A1 hasta la última celda)
    min_col = 1
    min_row = 1
    max_col = ws.max_column
    max_row = ws.max_row
    table_ref = f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:{openpyxl.utils.get_column_letter(max_col)}{max_row}"

    # Crear tabla
    tab = Table(displayName="PrediccionesTable", ref=table_ref)

    # Estilo de tabla
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )
    tab.tableStyleInfo = style
    
    # Añadir tabla a la hoja
    ws.add_table(tab)

    # Guardar Excel final
    wb.save(excel_output)

    print(f"\n[OK] Resultados guardados en Excel: {excel_output}")

if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Inferencia con modelo MLP (usa preprocessor.joblib y Model_V1.h5)")
    p.add_argument("csv", help="Archivo CSV de entrada (está en la misma carpeta)")
    p.add_argument("--preprocessor", default="preprocessor.joblib", help="Archivo joblib del preprocessor (default: preprocessor.joblib)")
    p.add_argument("--model", default="Model_V1.h5", help="Archivo del modelo Keras (default: Model_V1.h5)")
    p.add_argument("--output", default="predictions_output.csv", help="CSV de salida con probabilidades y predicción")
    p.add_argument("--default-threshold", type=float, default=0.5, dest="default_threshold", help="Umbral por defecto si no hay labels para optimizar (default 0.5)")
    args = p.parse_args()

    try:
        main(args)
    except Exception as exc:
        print(f"[ERROR FATAL] {exc}")
        sys.exit(2)